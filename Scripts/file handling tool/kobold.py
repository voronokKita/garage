#!python3
"""
I have a very important project for which I came up with a special file naming structure.
Each file has a date, a four-digit counter, (optional 1-3 letters of the category,) and a name.
Like:
2021.01.12.0000.file, 2021.01.12.0001.file, 2021.01.12.0002.file, 2021.01.12.0003.b.file

I found this clear naming structure very nice. 
I can always change the directories of the project or even put all the files in one directory, 
and they all line up in a careful order, no two files can conflict. 

Since the project is huge, it is impossible for me to keep it in order myself. 
I wrote this tool to automatically check all files, regardless of the directory in which they are located. 
The script points me to files with incorrect names, 
brings the names to the standard and fixes the counters,
in cases where they are messed up or the files have been deleted. 

v2 2021.08 all names to snake_case, ensure that all files have extensions
v3 make exel spreadsheet of all files
"""
import re
import sys
import shutil
import pathlib
import openpyxl
from time import sleep
from send2trash import send2trash
from collections import defaultdict
from openpyxl.styles import NamedStyle, Font, Alignment


STEP = 0.1
SPEED = 2

SIZE = 11
FONT = "Arial"
SPREADSHEET_NAME = "project_table.xlsx"

""" Project structure """
folder = pathlib.Path(__file__).resolve().parent
PROJECT = [
    pathlib.Path(folder, "1 dir"),
    pathlib.Path(folder, "2 dir"),
    pathlib.Path(folder, "3 dir"),
    pathlib.Path(folder, "4 dir"),
    pathlib.Path(folder, "5 dir"),
    pathlib.Path(folder, "6 dir"),
]
OLD_FILES = pathlib.Path(folder, "OLD")

PATTERN = re.compile(r'''(
        ((\d{4}\.\d\d\.\d\d)\.)     # date
        (\d{4})                     # counter 04d int
        (\.([A-Z]{1,3}))?           # category 1-3 char (optional)
        (\.([^.]{5,}))              # filename 5+ char
        (\.[A-Z]{2,4})              # extention 2-4 char
    )''', re.VERBOSE | re.I)
""" Result in 8 parts:
[('2021.08.03.0000.b.new_name.txt', '2021.08.03.', '2021.08.03', '0000', '.b', 'b', '.new_name', 'new_name', '.txt')]
[('2021.08.03.0001.new_name.txt', '2021.08.03.', '2021.08.03', '0000', '', '', '.new_name', 'new_name', '.txt')] """


class FilesLostError(Exception):
    pass


def main():
    sleep(STEP)
    initial_amount = len(load_files())
    print(f"project has {initial_amount} files")
    
    if bad_names_structure():
        sys.exit(1)
        
    OLD_FILES.mkdir()

    lower_snake_case()
    check_quantity(initial_amount)
    
    incorrect_counters = check_bad_counters()
    if incorrect_counters:
        recount(incorrect_counters)
        check_quantity(initial_amount)

    sleep(SPEED)
    print("clear")
    
    make_spreadsheet()
    
    send2trash(str(OLD_FILES))
    print("done.")
    sys.exit(0)
        

def load_files():
    """ Load paths of files 
        return: [list] of files """
    project_content = []    
    for directory in PROJECT:
        assert directory.exists(), "STOP. project structure has changed"
        project_content += [item for item in directory.rglob('*') if item.is_file()]
    return project_content


def check_quantity(initial_amount):
    files = len(load_files())
    if files != initial_amount:
        raise FilesLostError(f"ERROR. number of files changed! \
                               was {initial_amount}, became {files}")


def bad_names_structure():
    """ Check files naming structure 
        return: bool """
    sleep(SPEED)
    errors = False
    for file in load_files():
        if not PATTERN.search(file.name):
            print("incorrect filename:\t", file)
            if not errors:
                errors = True
            sleep(STEP)
    
    print("incorrect names.") if errors else print("OK names structure")
    return errors


def lower_snake_case():
    """ Check filenames case """
    sleep(SPEED)
    errors = 0
    for file in load_files():

        if " " not in file.name and file.name.islower():
            continue
        else:
            print("to lower_snake_case:\t", file.name)

            old_file = pathlib.Path(OLD_FILES, f"{file.name} (lower_snake_case)")
            shutil.copy(file, old_file)

            lower_case = file.name.lower()
            parts = PATTERN.findall(lower_case)[0]
            file_date = parts[1]
            counter = parts[3]
            category = parts[4]
            old_name = parts[7]
            snake_case = "." + "_".join(old_name.strip().split(" "))
            extension = parts[8]
            new_name = file_date + counter + category + snake_case + extension

            if pathlib.Path(file.parent, new_name).exists():  # little paranoia
                new_name = f"{errors}!{new_name}"

            file.rename(pathlib.Path(file.parent, new_name))

            errors += 1
            sleep(STEP)
        
    print(f"{errors} files renamed") if errors else print("OK lower_snake_case")


def check_bad_counters():
    """ Sort files on name and group them on date,
        then check files counters.
        return: {dict} of dates with incorrect counters """
    sleep(SPEED)
    # fill dict of dates
    dates_dictionary = defaultdict(list)
    for file in sorted(load_files(), key=lambda f: f.name):
        date = PATTERN.findall(file.name)[0][2]
        dates_dictionary[date] += [file]

    # fill dict of dates with incorrect counters
    errors = 0
    incorrect_counters = {}
    for date in dates_dictionary:
        for i, file in enumerate(dates_dictionary[date]):
            counter = PATTERN.findall(file.name)[0][3]
            if i != int(counter):
                print(date, "incorrect counters")
                incorrect_counters[date] = dates_dictionary[date]
                errors += 1
                sleep(STEP)
                break

    print(f"counters incorrect for {errors} dates") if errors else print("OK counters")
    return incorrect_counters


def recount(incorrect_counters):
    """ Overwrite counters;
        duplicate the process to avoid conflicts """
    sleep(SPEED)
    print("recount:")

    errors = 0
    renamed = []
    for date in incorrect_counters:
        for i, file in enumerate(incorrect_counters[date]):
            old_file = pathlib.Path(OLD_FILES, f"{file.name} (old)")
            shutil.copy(file, old_file)

            parts = PATTERN.findall(file.name)[0]
            file_date = parts[1]
            counter = f"{i:04d}"
            category = parts[4]
            filename = parts[6]
            extension = parts[8]
            new_name = file_date + counter + category + filename + extension
            print(new_name)

            file = file.rename(pathlib.Path(file.parent, f"{new_name} (new)"))
            renamed.append(file)
            
            errors += 1
            sleep(STEP)
    
    for file in renamed:
        new_name = file.name.replace(" (new)", "")
        file.rename(pathlib.Path(file.parent, new_name))

    print(errors, "files renamed")


def make_spreadsheet():
    """
    Create a table by dividing all files into two groups, one after the other:
            id  date        category    name    path
    head    0   yyyy.mm.dd  gd          text1   /.
    ...
    bottom  99  yyyy.mm.dd              text99  /.
    """
    sleep(SPEED)
    print("updating spreadsheet")
    book = style_spreadsheet()
    sheet = book["Main"]
    styles = ["", "numeric", "date", "category", "name", "path"]
    
    # make two list's of files according to category
    id_ = 0
    row = 2    
    for half in ("head", "bottom"):
        files = sorted(load_files(), key=lambda f: f.name, reverse=True)

        for file in files:
            parts = PATTERN.findall(file.name)[0]
            category = parts[5]
            
            if half == "head" and not category:
                continue
            elif half == "bottom" and category:
                continue

            date = parts[2]            
            name = parts[7]
            file_path = str(file.parent)
            line = [id_, date, category, name, file_path]

            for column, value in enumerate(line, start=1):                
                if column == 4:
                    if not category or category == "d":
                        style = styles[column]
                    elif category == "gd":
                        style = "gd"
                    elif category == "ld":
                        style = "ld"
                    elif category == "gld":
                        style = "gld"
                    elif category == "fld":
                        style = "fld"
                else:                
                    style = styles[column]

                cell = sheet.cell(row=row, column=column)
                cell.style = style
                cell.value = value

            id_ += 1
            row += 1

        else:
            if half == "head":
                sheet.append(["", "", "", "", ""])
                sheet.append(["", "", "", "", ""])
                sheet.append(["", "", "", "", ""])
                row += 3

    old = pathlib.Path(folder, SPREADSHEET_NAME)
    if old.exists():
        old.rename(pathlib.Path(OLD_FILES, f"{old.name} (old)"))
    
    book.save(SPREADSHEET_NAME)


def style_spreadsheet():
    book = openpyxl.Workbook()
    book.create_sheet(title="Main", index=0)
    sheet = book["Main"]

    # apply width values
    sheet.column_dimensions['A'].width = 8  # id
    sheet.column_dimensions['B'].width = 12  # date
    sheet.column_dimensions['C'].width = 5  # category
    sheet.column_dimensions['D'].width = 42  # filename
    sheet.column_dimensions['E'].width = 70  # path

    # create named styles
    style = NamedStyle(name="numeric")
    style.font = Font(name=FONT, size=SIZE-2, color='000000')
    style.alignment = Alignment(vertical='center')
    book.add_named_style(style)

    style = NamedStyle(name="date", number_format='YYYY.MM.DD')
    style.font = Font(name=FONT, size=SIZE, color='000000')
    style.alignment = Alignment(horizontal='center', vertical='center')
    book.add_named_style(style)

    style = NamedStyle(name="category")
    style.font = Font(name=FONT, size=SIZE, bold=True, color='000000')
    style.alignment = Alignment(horizontal='center', vertical='center')
    book.add_named_style(style)

    style = NamedStyle(name="gd")
    style.font = Font(name=FONT, size=SIZE, bold=True, color='ff7f50')
    style.alignment = Alignment(vertical='center')
    book.add_named_style(style)

    style = NamedStyle(name="ld")
    style.font = Font(name=FONT, size=SIZE, bold=True, color='9400d3')
    style.alignment = Alignment(vertical='center')
    book.add_named_style(style)

    style = NamedStyle(name="gld")
    style.font = Font(name=FONT, size=SIZE, bold=True, color='ff00ff')
    style.alignment = Alignment(vertical='center')
    book.add_named_style(style)

    style = NamedStyle(name="fld")
    style.font = Font(name=FONT, size=SIZE, bold=True, color='0000ff')
    style.alignment = Alignment(vertical='center')
    book.add_named_style(style)

    style = NamedStyle(name="name")
    style.font = Font(name=FONT, size=SIZE, color='000000')
    style.alignment = Alignment(vertical='center')
    book.add_named_style(style)
    
    style = NamedStyle(name="path")
    style.font = Font(name=FONT, size=SIZE-2, color='000000')
    style.alignment = Alignment(vertical='center')
    book.add_named_style(style)

    sheet.append(["id", "date", "char", "name", "path"])
    for cell in sheet['1:1']:
        cell.style = "category"
    sheet.freeze_panes = 'A2'

    return book


if __name__ == "__main__":
    main()
