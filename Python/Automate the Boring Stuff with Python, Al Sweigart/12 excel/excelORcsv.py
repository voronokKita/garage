#! python3
""" spring 2021 Convert csv to excel and vice versa. """
import os
import re
import sys
import csv
import openpyxl
from send2trash import send2trash
from openpyxl.styles import NamedStyle, Font
from openpyxl.utils.cell import _get_column_letter

if os.path.exists("generatedata.com.csv"):
    # open excel object;
    book = openpyxl.Workbook()
    book.create_sheet(title="Main", index=0)
    sheet = book["Main"]

    # create named styles;
    new_stile = NamedStyle(name="date", number_format='DD.MM.YY')
    new_stile.font = Font(name='Arial', size=11, color='000000')
    book.add_named_style(new_stile)
    new_stile = NamedStyle(name="currency")
    new_stile.font = Font(name='Courier New', size=10.5, color='61380B')
    book.add_named_style(new_stile)
    new_stile = NamedStyle(name="numeric")
    new_stile.font = Font(name='Courier New', size=10, color='000000')
    book.add_named_style(new_stile)
    new_stile = NamedStyle(name="text")
    new_stile.font = Font(name='Philosopher', size=12, color='0A1B2A')
    book.add_named_style(new_stile)

    with open("generatedata.com.csv") as inputfile:
        reader = csv.DictReader(inputfile)
        # write header;
        sheet.append(reader.fieldnames)

        column_widths = []
        start_row = 2
        start_column = 1
        for r, row in enumerate(reader, start_row):
            for c, key in enumerate(row, start_column):
                cell = sheet.cell(row=r, column=c)
                # save the widest column width's;
                index = c - 1
                cell_width = len(row[key])
                if len(column_widths) < c:
                    column_widths.append(cell_width)
                elif column_widths[index] < cell_width:
                    column_widths[index] = cell_width
                # define style;
                if re.search(r'(\d\d).(\d\d).(\d\d)', row[key]) is not None:
                    style = "date"
                elif "€" in row[key]:
                    style = "currency"
                else:
                    try:
                        num = int(row[key])
                        style = "numeric"
                    except ValueError:
                        style = "text"
                # write cell;
                cell.style = style
                cell.value = row[key]

        # apply width values;
        for c, cell_width in enumerate(column_widths, start_column):
            column = _get_column_letter(c)
            sheet.column_dimensions[column].width = cell_width

    # create excel file and remove csv.
    try:
        book.save("generatedata.com.xlsx")
    except OSError as error:
        print("ERROR saving excel file.", error)
        sys.exit(1)
    else:
        send2trash("generatedata.com.csv")
        sys.exit(0)

elif os.path.exists("generatedata.com.xlsx"):
    # load excel file;
    book = openpyxl.load_workbook(filename="generatedata.com.xlsx")
    sheet = book.active

    # write csv and remove xlsx;
    start = 1
    max_rows = sheet.max_row + 1
    max_column = sheet.max_column + 1
    try:
        with open("generatedata.com.csv", 'w') as outputfile:
            writer = csv.writer(outputfile)
            for r in range(start, max_rows):
                row = []
                for c in range(start, max_column):
                    cell = sheet.cell(row=r, column=c)
                    value = cell.value
                    row.append(value)
                writer.writerow(row)
    except OSError as error:
        print("ERROR write csv.", error)
        sys.exit(2)
    else:
        send2trash("generatedata.com.xlsx")
        sys.exit(0)

else:
    print("generatedata.com file non found.")
    sys.exit(3)
